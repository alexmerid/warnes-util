import openpyxl
import pymysql

# Configuración de la conexión
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': 'root',
    'database': 'warnes',
    'cursorclass': pymysql.cursors.DictCursor
}

conexion = pymysql.connect(**DB_CONFIG)
cursor = conexion.cursor()

# nomarch = "/media/alexander/Unidad_E/Warnes/Ejecucion/2025-06/000.xlsx"
nomarch = input("Ingrese el nombre del archivo Excel (con ruta completa): ")
# El archivo Excel tiene la siguiente estructura:
# id_poste, id_luminaria, codigo, id_via, fecha_cambio, lum_ant
wb = openpyxl.load_workbook(nomarch)
hoja = wb.active

print("----- Inicio del proceso -----")
# Verficar si los Codigos de luminarias no existen en la base de datos
existentes = []
observaciones = []
for fila in hoja.iter_rows(min_row=2, values_only=True):
    if fila[0] is None:
        continue
    cursor.execute(
        "select count(*) as conteo from poste_luminaria where codigo = %s", (fila[2],))
    resultado = cursor.fetchone()
    if resultado['conteo'] > 0:
        existentes.append(fila[2])
    cursor.execute(
        "SELECT id_poste FROM observacion WHERE id_poste  = %s", (fila[0],))
    resultado = cursor.fetchone()
    if resultado:
        observaciones.append((fila[0], f"{fila[4]:%Y-%m-%d}"))
if existentes:
    print("Los siguientes códigos ya existen en la base de datos:")
    print(existentes)
    raise ValueError(
        "Se encontraron Códigos duplicados. Proceso detenido.")
if observaciones:
    print("Los siguientes postes tienen observaciones que deben finalizarse:")
    print(observaciones)

# Leer el archivo Excel y procesar los datos
c = 0
id_ant = 0
for fila in hoja.iter_rows(min_row=2, values_only=True):
    if fila[0] is None:
        continue
    cursor.execute(
        "select id from poste_luminaria where id_poste = %s and id_luminaria = %s and fecha_desinst is null and codigo is null", (fila[0], fila[5]))
    resultado = cursor.fetchone()
    if resultado:
        cursor.execute(
            "update poste_luminaria set fecha_desinst=%s where id = %s", (f"{fila[4]:%Y-%m-%d}", resultado['id']))
        cursor.execute("insert into poste_luminaria (id_poste, id_luminaria, estado, fecha_inst, codigo, reemp) values(%s, %s, 1, %s, %s, %s)",
                       (fila[0], fila[1], f"{fila[4]:%Y-%m-%d}", fila[2], resultado['id']))
        if id_ant != fila[0] and fila[3] != '':
            cursor.execute(
                "update poste set id_via=%s where id=%s", (fila[3], fila[0]))
        conexion.commit()
        c += 1
    else:
        print(
            f"No se encontró poste_luminaria para poste {fila[0]} y luminaria {fila[2]}.")
    id_ant = fila[0]
print(f"Se actualizaron {c} filas")
print("----- Fin del proceso -----")

cursor.close()
conexion.close()
