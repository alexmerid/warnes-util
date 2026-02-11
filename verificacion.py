import openpyxl
import pymysql

# Configuración de la conexión
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': 'root',
    'database': 'warnes',
    'cursorclass': pymysql.cursors.DictCursor
}

conexion = pymysql.connect(**DB_CONFIG)
cursor = conexion.cursor()

nomarch = input("Ingrese el nombre del archivo Excel (con ruta completa): ")

wb = openpyxl.load_workbook(nomarch)
hoja = wb.active
# print("Filas = ", hoja.max_row)
print("\n")
i = 1
c = 0
for fila in hoja.iter_rows(min_row=7, values_only=True):
    if fila[3] is not None:
        cursor.execute(
            "SELECT id_poste FROM poste_luminaria WHERE codigo = %s", (fila[3],))
        resultado = cursor.fetchone()
        if resultado:
            if resultado['id_poste'] == fila[0]:
                print(f"{i:>4}. {fila[0]:>6} {fila[3]:>8} - OK")
                c += 1
            else:
                print(
                    f"{i:>4}. {fila[0] if fila[0] is not None else 'N/A':>6} {fila[3]:>8} -> {resultado['id_poste']:>6} - ERROR")
        else:
            print(f"{i:>4}. {fila[0]:>6} {fila[3]:>8} - NO EXISTE")
        i += 1
print(f"\nCorrectos: {c} de {i-1}")
