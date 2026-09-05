# Programa para la etapa de ampliación.
import openpyxl
import pymysql

# Configuración de la conexión
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': 'root',
    'database': 'warnes',
    'cursorclass': pymysql.cursors.DictCursor
}

conexion = pymysql.connect(**DB_CONFIG)
cursor = conexion.cursor()

# nomarch = "/home/alexander/Unidad_D/Warnes/Ejecucion 2026/2026-08/2026-08-26/tmp.xlsx"
nomarch = input("Ingrese el nombre del archivo Excel (con ruta completa): ")
# El archivo Excel tiene la siguiente estructura:
# id_poste, id_luminaria, codigo, id_via, fecha_inst, id_referencia, latitud, longitud, obs
wb = openpyxl.load_workbook(nomarch)
hoja = wb.active

print("\n------------------------------------------------------------")
# Verficar si los Codigos de luminarias no existen en la base de datos
existentes = []
observaciones = []
id_ant = 0
postes_obs = []
for fila in hoja.iter_rows(min_row=2, values_only=True):
    if fila[0] is None:
        continue
    cursor.execute(
        "select count(*) as conteo from poste_luminaria where codigo = %s", (fila[2],))
    resultado = cursor.fetchone()
    if resultado['conteo'] > 0:
        existentes.append(fila[2])
    cursor.execute(
        "SELECT id_poste FROM observacion WHERE id_poste  = %s", (fila[0],))
    resultado = cursor.fetchone()
    if resultado:
        observaciones.append((fila[0], f"{fila[4]:%Y-%m-%d}"))
    if id_ant == fila[0] and fila[8]:
        postes_obs.append(fila[0])
    id_ant = fila[0]
if existentes:
    print("Los siguientes códigos ya existen en la base de datos:")
    print(existentes)
    raise ValueError(
        "Se encontraron Códigos duplicados. Proceso detenido.")
if observaciones:
    print("Los siguientes postes ya tienen observaciones:")
    print(observaciones)
    continuar = input(
        "¿Desea continuar con la inserción de los datos? (s/n): ")
    if continuar.lower() != 's':
        raise ValueError("Proceso detenido por el usuario.")
if postes_obs:
    print("Los siguientes postes tienen más de una observacion:")
    print(postes_obs)
    continuar = input(
        "¿Desea continuar con la inserción de los datos? (s/n): ")
    if continuar.lower() != 's':
        raise ValueError("Proceso detenido por el usuario.")

# Leer el archivo Excel y procesar los datos
p = 0
l = 0
o = 0
for fila in hoja.iter_rows(min_row=2, values_only=True):
    if fila[0] is None:
        continue
    # Verificar si el poste ya existe en la base de datos
    cursor.execute(
        "SELECT id FROM poste WHERE id = %s", (fila[0], ))
    resultado = cursor.fetchone()
    if resultado:
        # Modificar id_referencia y id_via del poste existente
        cursor.execute("UPDATE poste SET id_referencia=%s, id_via=%s WHERE id=%s",
                       (fila[5], fila[3], fila[0]))
    else:
        # Insertar un nuevo poste en la base de datos
        cursor.execute("INSERT INTO poste (id, latitud, longitud, id_referencia, id_via) VALUES(%s, %s, %s, %s, %s)",
                       (fila[0], fila[6], {fila[7]}, fila[5], {fila[3]}))
        p += 1
    # Insertar la luminaria en la tabla poste_luminaria
    cursor.execute("INSERT INTO poste_luminaria(id_poste, id_luminaria, estado, fecha_inst, codigo) VALUES(%s, %s, 1, %s, %s)",
                   (fila[0], fila[1], f"{fila[4]:%Y-%m-%d}", fila[2]))

    # Si la observación no está en blanco, insertar la observación en la tabla observacion
    if fila[8]:
        cursor.execute("INSERT INTO observacion(id_poste, fecha_obs, descripcion) VALUES(%s, %s, %s)",
                       (fila[0], f"{fila[4]:%Y-%m-%d}", fila[8]))
        o += 1
    conexion.commit()
    l += 1

print(f"Se insertaron {p} postes, {l} luminarias y {o} observaciones.")
print("------------------------------------------------------------\n")

cursor.close()
conexion.close()
